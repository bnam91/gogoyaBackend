from pymongo import MongoClient
from pymongo.server_api import ServerApi
from datetime import datetime, timedelta
import json
from datetime import timezone
from collections import defaultdict
import pandas as pd  # pandas 추가
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import AutoFilter

# MongoDB 연결 설정
uri = "mongodb+srv://coq3820:JmbIOcaEOrvkpQo1@cluster0.qj1ty.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
client = MongoClient(uri, server_api=ServerApi('1'), 
                    connectTimeoutMS=60000,  # 연결 타임아웃을 60초로 증가
                    socketTimeoutMS=60000)   # 소켓 타임아웃도 60초로 증가

try:
    # 연결 확인
    client.admin.command('ping')
    print("MongoDB 연결 성공!")
    
    # 데이터베이스와 컬렉션 선택
    db = client['insta09_database']
    collection = db['02_main_influencer_data']
    
    # 사용자로부터 인플루언서 username 입력 받기
    target_username = input("분석할 인플루언서의 username을 입력하세요: ")
    
    # 사용자로부터 날짜 입력 받기
    while True:
        try:
            days_ago = int(input("몇 일 전부터의 데이터를 분석할까요? (숫자만 입력): "))
            if days_ago > 0:
                break
            else:
                print("1 이상의 숫자를 입력해주세요.")
        except ValueError:
            print("올바른 숫자를 입력해주세요.")
    
    # 입력받은 username이 데이터베이스에 존재하는지 확인
    user_exists = collection.find_one({"username": target_username})
    if not user_exists:
        raise Exception(f"'{target_username}' 사용자를 찾을 수 없습니다. 올바른 username을 입력해주세요.")
    
    # 입력받은 날짜만큼 이전 날짜 계산 (UTC 기준)
    ninety_days_ago = datetime.now(timezone.utc) - timedelta(days=days_ago)
    
    # 기본 쿼리로 데이터 가져오기
    user_data = collection.find_one(
        {"username": target_username},
        {"brand": 1, "clean_name": 1, "reels_views(15)": 1}
    )
    
    # clean_name 필드 출력
    user_clean_name = user_data.get('clean_name', '')
    user_reels_views = user_data.get('reels_views(15)')
    print("\nclean_name 필드:", user_clean_name)
    print("Reels Views (15):", user_reels_views if user_reels_views is not None else 'N/A')
    
    # 제외할 브랜드명 리스트
    excluded_brands = ['확인필요', 'n', 'N', 'N/A', '복합상품']
    
    # 대상 인플루언서가 진행한 브랜드 목록 수집
    target_brands = set()
    if user_data and 'brand' in user_data:
        for brand in user_data['brand']:
            brand_name = brand.get('name', '')
            clean_name = brand.get('clean_name', '')
            if (brand_name not in excluded_brands and 
                target_username not in clean_name.lower() and
                target_username not in brand_name.lower() and
                user_clean_name.lower() not in brand_name.lower()):  # 타겟 유저의 clean_name이 브랜드명에 포함되지 않은 경우만
                for product in brand.get('products', []):
                    try:
                        mentioned_date = datetime.fromisoformat(product['mentioned_date'].replace('Z', '+00:00'))
                        if mentioned_date >= ninety_days_ago:
                            target_brands.add(brand_name)
                            break
                    except ValueError:
                        continue
    
    # 결과 정리
    brand_dict = {}  # brand_name을 키로 사용하여 최신 데이터만 저장
    if user_data and 'brand' in user_data:
        for brand in user_data['brand']:
            brand_name = brand.get('name', '')
            clean_name = brand.get('clean_name', '')
            
            # 제외 조건 검사
            if (brand_name not in excluded_brands and
                target_username not in clean_name.lower() and
                target_username not in brand_name.lower() and
                user_clean_name.lower() not in brand_name.lower()):  # 타겟 유저의 clean_name이 브랜드명에 포함되지 않은 경우만
                for product in brand.get('products', []):
                    # UTC timezone을 가진 datetime으로 변환
                    mentioned_date = datetime.fromisoformat(product['mentioned_date'].replace('Z', '+00:00'))
                    if mentioned_date >= ninety_days_ago:
                        brand_info = {
                            'brand_name': brand_name,
                            'clean_name': clean_name,
                            'product_item': product['item'],
                            'product_category': product['category'],
                            'item_feed_link': product['item_feed_link'],
                            'mentioned_date': product['mentioned_date'],
                            'mentioned_date_obj': mentioned_date  # 정렬을 위해 datetime 객체 추가
                        }
                        
                        # 해당 브랜드의 기존 데이터가 없거나, 현재 데이터가 더 최신인 경우 업데이트
                        if brand_name not in brand_dict or brand_dict[brand_name]['mentioned_date_obj'] < mentioned_date:
                            brand_dict[brand_name] = brand_info
    
    # 딕셔너리를 리스트로 변환하고 mentioned_date_obj 제거
    brand_list = []
    for brand_info in brand_dict.values():
        del brand_info['mentioned_date_obj']
        brand_list.append(brand_info)
    
    # JSON 형태로 출력
    print(f"\n=== {target_username}이 진행한 브랜드 목록 ===")
    print(json.dumps(brand_list, indent=2, default=str, ensure_ascii=False))
    
    # 각 브랜드별로 진행한 다른 인플루언서 찾기
    print("\n=== 각 브랜드를 진행한 다른 인플루언서 목록 ===")
    
    # 모든 관련 인플루언서 목록 수집
    all_influencers = set()
    brand_to_influencers = defaultdict(list)  # 브랜드별 인플루언서 정보를 저장할 딕셔너리
    
    # 먼저 대상 인플루언서의 브랜드 목록 수집
    target_brands = set()
    for brand_info in brand_list:
        target_brands.add(brand_info['brand_name'])
    
    # 다른 인플루언서들의 브랜드 정보 수집
    cursor = collection.find(
        {
            "username": {"$ne": target_username},
            "brand": {
                "$elemMatch": {
                    "name": {"$in": list(target_brands)}  # 대상 인플루언서의 브랜드만 찾기
                }
            }
        }
    )
    
    for inf_data in cursor:
        username = inf_data['username']
        clean_name = inf_data.get('clean_name', '')
        reels_views = inf_data.get('reels_views(15)', 'N/A')
        
        if 'brand' in inf_data:
            for brand in inf_data['brand']:
                brand_name = brand.get('name', '')
                
                # 대상 인플루언서의 브랜드인 경우에만 처리
                if brand_name in target_brands:
                    for product in brand.get('products', []):
                        try:
                            mentioned_date = datetime.fromisoformat(product['mentioned_date'].replace('Z', '+00:00'))
                            if mentioned_date >= ninety_days_ago:
                                brand_to_influencers[brand_name].append({
                                    'username': username,
                                    'clean_name': clean_name,
                                    'product_item': product['item'],
                                    'product_category': product['category'],
                                    'mentioned_date': product['mentioned_date'],
                                    'item_feed_link': product.get('item_feed_link', ''),
                                    'reels_views': reels_views
                                })
                                all_influencers.add((username, clean_name))
                                break
                        except ValueError:
                            continue
    
    # 각 인플루언서별로 최근 90일 내 진행한 브랜드 출력
    print("\n=== 각 인플루언서별 최근 90일 내 진행한 브랜드 목록 ===")
    
    for username, clean_name in sorted(all_influencers):
        print(f"\n인플루언서: {username} (clean_name: {clean_name})")
        print("진행한 브랜드 목록:")
        
        # 해당 인플루언서의 브랜드 데이터 가져오기
        inf_data = collection.find_one(
            {"username": username},
            {
                "brand": 1,
                "reels_views(15)": 1
            }
        )
        
        if inf_data and 'brand' in inf_data:
            recent_brands = []
            for brand in inf_data['brand']:
                brand_name = brand.get('name', '')
                if brand_name not in excluded_brands:
                    for product in brand.get('products', []):
                        try:
                            mentioned_date = datetime.fromisoformat(product['mentioned_date'].replace('Z', '+00:00'))
                            if mentioned_date >= ninety_days_ago:
                                brand_info = {
                                    'brand_name': brand_name,
                                    'product_item': product['item'],
                                    'product_category': product['category'],
                                    'mentioned_date': product['mentioned_date']
                                }
                                recent_brands.append(brand_info)
                                # 브랜드별 인플루언서 정보 저장 (대상 인플루언서가 진행하지 않은 브랜드만)
                                if brand_name not in target_brands:
                                    # 인플루언서의 reels_views 가져오기
                                    reels_views = inf_data.get('reels_views(15)')
                                    brand_to_influencers[brand_name].append({
                                        'username': username,
                                        'clean_name': clean_name,
                                        'product_category': product['category'],
                                        'mentioned_date': product['mentioned_date'],
                                        'product_item': product['item'],
                                        'item_feed_link': product.get('item_feed_link', ''),
                                        'reels_views': reels_views if reels_views is not None else 'N/A'
                                    })
                                break  # 각 브랜드당 최신 제품 하나만 포함
                        except ValueError:
                            continue
            
            if recent_brands:
                for brand in recent_brands:
                    print(f"- {brand['brand_name']}")
                    print(f"  제품: {brand['product_item']}")
                    print(f"  언급일자: {brand['mentioned_date']}")
            else:
                print("- 최근 90일 내 진행한 브랜드가 없습니다.")
        else:
            print("- 브랜드 데이터가 없습니다.")
    
    # 브랜드별 인플루언서 목록 출력 (대상 인플루언서 제외)
    print(f"\n=== 브랜드별 진행 인플루언서 종합 정리 ({target_username} 제외) ===")
    
    # Excel 데이터를 위한 리스트 생성
    excel_data = []
    
    for brand_name, influencers in sorted(brand_to_influencers.items()):
        for inf in sorted(influencers, key=lambda x: x['mentioned_date'], reverse=True):
            # Excel 데이터에 추가
            excel_data.append({
                '브랜드명': brand_name,
                '인플루언서': inf['username'],
                'Clean Name': inf['clean_name'],
                '제품': inf['product_item'],
                '아이템 카테고리': inf.get('product_category', ''),  # category -> product_category로 수정
                '언급일자': inf['mentioned_date'],
                '링크': inf['item_feed_link'],
                'Reels Views (15)': inf['reels_views']
            })
    
    # DataFrame 생성 및 Excel 파일로 저장
    if excel_data:
        df_others = pd.DataFrame(excel_data)
        # 칼럼 순서 재정렬
        df_others = df_others[['브랜드명', '제품', '아이템 카테고리', 'Clean Name', '인플루언서', '언급일자', '링크', 'Reels Views (15)']]
        # 언급일자 기준으로 최신순 정렬
        df_others = df_others.sort_values('언급일자', ascending=False)
        
        # 대상 인플루언서의 브랜드 데이터를 위한 리스트 생성
        target_data = []
        target_brands_dict = {}  # 브랜드별 상세 정보를 저장하기 위한 딕셔너리
        
        if user_data and 'brand' in user_data:
            for brand in user_data['brand']:
                brand_name = brand.get('name', '')
                clean_name = brand.get('clean_name', '')
                
                # 제외 조건 검사
                if (brand_name not in excluded_brands and
                    target_username not in clean_name.lower() and
                    target_username not in brand_name.lower() and
                    user_clean_name.lower() not in brand_name.lower()):  # 타겟 유저의 clean_name이 브랜드명에 포함되지 않은 경우만
                    for product in brand.get('products', []):
                        try:
                            mentioned_date = datetime.fromisoformat(product['mentioned_date'].replace('Z', '+00:00'))
                            if mentioned_date >= ninety_days_ago:
                                brand_info = {
                                    '브랜드명': brand_name,
                                    '인플루언서': target_username,
                                    '제품': product['item'],
                                    '아이템 카테고리': product['category'],
                                    'Clean Name': user_clean_name,
                                    '언급일자': product['mentioned_date'],
                                    '링크': product.get('item_feed_link', ''),
                                    'Reels Views (15)': user_reels_views if user_reels_views is not None else 'N/A'
                                }
                                target_data.append(brand_info)
                                target_brands_dict[brand_name] = brand_info
                                break
                        except ValueError:
                            continue
        
        df_target = pd.DataFrame(target_data)
        # 칼럼 순서 재정렬
        df_target = df_target[['브랜드명', '제품', '아이템 카테고리', 'Clean Name', '인플루언서', '언급일자', '링크', 'Reels Views (15)']]
        # 언급일자 기준으로 최신순 정렬
        df_target = df_target.sort_values('언급일자', ascending=False)
        
        # 겹치는 브랜드 정보 생성
        overlapping_data = []
        target_brands = set(target_brands_dict.keys())
        print(f"\n{target_username}의 브랜드 목록:", sorted(target_brands))
        
        # brand_to_influencers 내용 출력
        print("\n=== 다른 인플루언서들이 진행한 브랜드 목록 ===")
        for brand_name, influencers in brand_to_influencers.items():
            print(f"\n브랜드: {brand_name}")
            print(f"진행한 인플루언서 수: {len(influencers)}명")
            for inf in influencers:
                print(f"- {inf['username']} (clean_name: {inf['clean_name']})")
                print(f"  제품: {inf['product_item']}")
                print(f"  언급일자: {inf['mentioned_date']}")
        
        # brand_to_influencers에서 대상 인플루언서의 브랜드와 매칭되는 정보 찾기
        print(f"\n=== {target_username} 브랜드와 매칭되는 정보 찾기 ===")
        for brand_name, influencers in brand_to_influencers.items():
            if brand_name in target_brands:
                print(f"\n브랜드 '{brand_name}' 매칭됨!")
                target_info = target_brands_dict[brand_name]
                for inf in influencers:
                    print(f"- 매칭된 인플루언서: {inf['username']}")
                    overlapping_data.append({
                        '인플루언서': inf['username'],
                        '인플루언서_Clean_Name': inf['clean_name'],
                        '진행_브랜드': brand_name,
                        '인플루언서_제품': inf['product_item'],
                        '아이템_카테고리': inf.get('product_category', ''),  # category -> product_category로 수정
                        '인플루언서_언급일자': inf['mentioned_date'],
                        '인플루언서_링크': inf['item_feed_link'],
                        '인플루언서_Reels_Views': inf['reels_views']
                    })
        
        # DataFrame 생성 및 정렬
        df_overlapping = pd.DataFrame(overlapping_data)
        if not df_overlapping.empty:
            # 인플루언서 username의 중복 횟수를 계산하여 정렬
            influencer_counts = df_overlapping['인플루언서'].value_counts()
            df_overlapping['중복_횟수'] = df_overlapping['인플루언서'].map(influencer_counts)
            
            # Reels Views를 숫자로 변환하는 함수
            def convert_reels_views(value):
                    try:
                        if isinstance(value, (int, float)):
                            return float(value)
                        if isinstance(value, str):
                            value = value.replace(',', '')
                            if 'K' in value:
                                return float(value.replace('K', '')) * 1000
                            elif 'M' in value:
                                return float(value.replace('M', '')) * 1000000
                            elif value.replace('.', '').isdigit():
                                return float(value)
                    except (ValueError, TypeError):
                        pass
                    return 0

            # Reels Views를 숫자로 변환하여 임시 컬럼 생성
            df_overlapping['Reels_Views_Numeric'] = df_overlapping['인플루언서_Reels_Views'].apply(convert_reels_views)
            
            # 중복 횟수 내림차순, Reels Views 내림차순으로 정렬
            df_overlapping = df_overlapping.sort_values(
                ['중복_횟수', 'Reels_Views_Numeric'], 
                ascending=[False, False]
            )
            
            # 임시 컬럼들 제거
            df_overlapping = df_overlapping.drop(['중복_횟수', 'Reels_Views_Numeric'], axis=1)
            print(f"\n총 {len(set(df_overlapping['인플루언서']))}명의 인플루언서가 {target_username}의 브랜드를 진행했습니다.")
        
        # Excel 파일로 저장
        excel_filename = f'brand_influencer_summary_{target_username}.xlsx'
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            if not df_target.empty:
                df_target.to_excel(writer, sheet_name='Sheet0', index=False)
                # Sheet0 칼럼 너비 설정 및 필터 추가
                worksheet = writer.sheets['Sheet0']
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Sheet0의 링크 컬럼(7번째 열)을 하이퍼링크로 설정
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):  # 헤더 제외
                    try:
                        link_cell = row[6]  # G열 (링크)
                        if link_cell.value and isinstance(link_cell.value, str) and (link_cell.value.startswith('http://') or link_cell.value.startswith('https://')):
                            link_cell.hyperlink = link_cell.value
                            link_cell.style = 'Hyperlink'
                    except Exception:
                        continue
                
                for column in worksheet.columns:
                    worksheet.column_dimensions[column[0].column_letter].width = 25

            if not df_others.empty:
                df_others.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # Sheet1 스타일 적용
                worksheet = writer.sheets['Sheet1']
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # 분홍색 배경 스타일 정의 (Sheet1용)
                pink_fill1 = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                
                # 현재 시간 기준 21일 전 계산
                now = datetime.now(timezone.utc)
                fourteen_days_ago = now - timedelta(days=21)
                
                # 브랜드별 최초 언급일자를 찾기 위한 딕셔너리
                brand_first_mentions = {}
                
                # df_others에서 브랜드별 최초 언급일자 찾기
                for _, row in df_others.iterrows():
                    brand_name = row['브랜드명']
                    try:
                        mentioned_date = datetime.fromisoformat(row['언급일자'].replace('Z', '+00:00'))
                        if brand_name not in brand_first_mentions or mentioned_date < brand_first_mentions[brand_name]:
                            brand_first_mentions[brand_name] = mentioned_date
                    except (ValueError, TypeError):
                        continue
                
                # 최근 14일 이내에 최초 언급된 브랜드 목록
                recent_brands = {brand: date for brand, date in brand_first_mentions.items() if date >= fourteen_days_ago}
                
                # Sheet1의 각 행을 순회하면서 최초 언급일자가 14일 이내인 브랜드 찾아 분홍색으로 칠하기
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):  # 헤더 제외
                    try:
                        brand_cell = row[0]  # A열 (브랜드명)
                        if brand_cell.value in recent_brands:
                            for cell in row:
                                cell.fill = pink_fill1
                                
                        # Sheet1의 링크 컬럼(7번째 열)을 하이퍼링크로 설정
                        link_cell = row[6]  # G열 (링크)
                        if link_cell.value and isinstance(link_cell.value, str) and (link_cell.value.startswith('http://') or link_cell.value.startswith('https://')):
                            link_cell.hyperlink = link_cell.value
                            link_cell.style = 'Hyperlink'
                    except (ValueError, TypeError):
                        continue
                
                # Sheet1 칼럼 너비 설정
                for column in worksheet.columns:
                    worksheet.column_dimensions[column[0].column_letter].width = 25

            if not df_overlapping.empty:
                df_overlapping.to_excel(writer, sheet_name='Sheet2', index=False)
                worksheet = writer.sheets['Sheet2']
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Sheet2의 링크 컬럼(인플루언서_링크)을 하이퍼링크로 설정
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):  # 헤더 제외
                    try:
                        link_cell = row[6]  # G열 (인플루언서_링크)
                        if link_cell.value and isinstance(link_cell.value, str) and (link_cell.value.startswith('http://') or link_cell.value.startswith('https://')):
                            link_cell.hyperlink = link_cell.value
                            link_cell.style = 'Hyperlink'
                    except Exception:
                        continue
                
                # Sheet4 생성 - 브랜드명 기준 중복 제거
                if not df_others.empty:
                    # 브랜드명 기준으로 첫 번째 등장하는 행만 유지
                    df_sheet4 = df_others.drop_duplicates(subset=['브랜드명'], keep='first')
                    df_sheet4.to_excel(writer, sheet_name='Sheet4', index=False)
                    
                    # Sheet4 칼럼 너비 설정 및 스타일 적용
                    worksheet4 = writer.sheets['Sheet4']
                    worksheet4.auto_filter.ref = worksheet4.dimensions
                    
                    # 분홍색 배경 스타일 정의
                    pink_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    
                    # 현재 시간 기준 14일 전 계산
                    now = datetime.now(timezone.utc)
                    fourteen_days_ago = now - timedelta(days=14)
                    
                    # 브랜드별 최초(가장 오래된) 언급일자 찾기
                    brand_first_mentions = {}
                    for _, row in df_others.iterrows():
                        brand_name = row['브랜드명']
                        try:
                            mentioned_date = datetime.fromisoformat(row['언급일자'].replace('Z', '+00:00'))
                            if brand_name not in brand_first_mentions or mentioned_date < brand_first_mentions[brand_name]:
                                brand_first_mentions[brand_name] = mentioned_date
                        except (ValueError, TypeError):
                            continue
                    
                    # Sheet4의 각 행을 순회하면서 최초 언급일자가 14일 이내인 브랜드 찾아 분홍색으로 칠하기
                    for row_idx, row in enumerate(worksheet4.iter_rows(min_row=2)):  # 헤더 제외
                        try:
                            brand_cell = row[0]  # A열 (브랜드명)
                            if brand_cell.value in brand_first_mentions:
                                first_mention_date = brand_first_mentions[brand_cell.value]
                                if first_mention_date >= fourteen_days_ago:
                                    for cell in row:
                                        cell.fill = pink_fill
                            
                            # Sheet4의 링크 컬럼(7번째 열)을 하이퍼링크로 설정
                            link_cell = row[6]  # G열 (링크)
                            if link_cell.value and isinstance(link_cell.value, str) and (link_cell.value.startswith('http://') or link_cell.value.startswith('https://')):
                                link_cell.hyperlink = link_cell.value
                                link_cell.style = 'Hyperlink'
                        except (ValueError, TypeError):
                            continue
                    
                    for column in worksheet4.columns:
                        worksheet4.column_dimensions[column[0].column_letter].width = 25
                    print("Sheet4: 브랜드명 기준 중복 제거된 목록")

                # Sheet5 생성 - 브랜드별 중복 횟수
                brand_counts = df_others['브랜드명'].value_counts().reset_index()
                brand_counts.columns = ['브랜드명', '진행_횟수']
                
                # 각 브랜드의 가장 최근 제품과 카테고리 정보 가져오기
                latest_products = df_others.sort_values('언급일자', ascending=False).drop_duplicates('브랜드명')[['브랜드명', '제품', '아이템 카테고리']]
                
                # 각 브랜드별 진행 인플루언서 목록 생성
                brand_influencers = {}
                for _, row in df_others.iterrows():
                    brand_name = row['브랜드명']
                    username = row['인플루언서']
                    if brand_name not in brand_influencers:
                        brand_influencers[brand_name] = set()
                    brand_influencers[brand_name].add(username)
                
                # 각 브랜드별 인플루언서 목록을 쉼표로 구분된 문자열로 변환
                brand_influencers_str = {brand: ', '.join(sorted(usernames)) for brand, usernames in brand_influencers.items()}
                
                # 브랜드 카운트와 최근 제품 정보 병합
                brand_counts = brand_counts.merge(latest_products, on='브랜드명', how='left')
                
                # 진행인원 정보 추가
                brand_counts['진행인원'] = brand_counts['브랜드명'].map(brand_influencers_str)
                
                # 컬럼 순서 재정렬
                brand_counts = brand_counts[['브랜드명', '제품', '아이템 카테고리', '진행_횟수', '진행인원']]
                
                # 진행 횟수 기준으로 내림차순 정렬
                brand_counts = brand_counts.sort_values('진행_횟수', ascending=False)
                
                # Sheet5에 저장
                brand_counts.to_excel(writer, sheet_name='Sheet5', index=False)
                
                # Sheet5 칼럼 너비 설정 및 필터 추가
                worksheet5 = writer.sheets['Sheet5']
                worksheet5.auto_filter.ref = worksheet5.dimensions

                # Sheet5에 최근 브랜드 하이라이트 적용 (독립적으로 구현)
                # 분홍색 배경 스타일 정의 (Sheet5용)
                pink_fill5 = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                
                # 현재 시간과 14일 전 기준 (Sheet5용)
                now5 = datetime.now(timezone.utc)
                fourteen_days_ago5 = now5 - timedelta(days=14)
                
                # 브랜드별 최초 언급일자 (독립적으로 다시 계산)
                brand_first_mentions5 = {}
                for _, row in df_others.iterrows():
                    brand_name = row['브랜드명']
                    try:
                        mentioned_date = datetime.fromisoformat(row['언급일자'].replace('Z', '+00:00'))
                        if brand_name not in brand_first_mentions5 or mentioned_date < brand_first_mentions5[brand_name]:
                            brand_first_mentions5[brand_name] = mentioned_date
                    except (ValueError, TypeError):
                        continue
                
                # Sheet5의 각 행에 분홍색 적용
                for row_idx, row in enumerate(worksheet5.iter_rows(min_row=2)):  # 헤더 제외
                    try:
                        brand_cell = row[0]  # A열 (브랜드명)
                        if brand_cell.value in brand_first_mentions5:
                            first_mention_date = brand_first_mentions5[brand_cell.value]
                            if first_mention_date >= fourteen_days_ago5:
                                for cell in row:
                                    cell.fill = pink_fill5
                    except (ValueError, TypeError):
                        continue
                
                for column in worksheet5.columns:
                    worksheet5.column_dimensions[column[0].column_letter].width = 25

                # Sheet3 생성 - 인플루언서별 통계
                influencer_counts = df_overlapping['인플루언서'].value_counts().reset_index()
                influencer_counts.columns = ['username', '브랜드_진행_횟수']
                
                # Clean Name 정보 추가
                influencer_clean_names = df_overlapping[['인플루언서', '인플루언서_Clean_Name']].drop_duplicates()
                influencer_counts['clean_name'] = influencer_counts['username'].map(
                    dict(zip(influencer_clean_names['인플루언서'], influencer_clean_names['인플루언서_Clean_Name']))
                )
                
                # Reels Views 정보 추가 (각 인플루언서의 최신 값 사용)
                influencer_reels = df_overlapping[['인플루언서', '인플루언서_Reels_Views']].drop_duplicates()
                influencer_counts['reels_views'] = influencer_counts['username'].map(
                    dict(zip(influencer_reels['인플루언서'], influencer_reels['인플루언서_Reels_Views']))
                )

                # 진행 브랜드 목록 생성
                brand_lists = df_overlapping.groupby('인플루언서')['진행_브랜드'].unique().apply(lambda x: ', '.join(sorted(x)))
                influencer_counts['brand_list'] = influencer_counts['username'].map(brand_lists)

                # profile_link 정보 가져오기
                profile_links = {}
                for username in influencer_counts['username']:
                    # MongoDB에서 해당 인플루언서의 profile_link 조회
                    user_data = collection.find_one(
                        {"username": username},
                        {"profile_link": 1}
                    )
                    if user_data and 'profile_link' in user_data:
                        profile_links[username] = user_data['profile_link']
                    else:
                        profile_links[username] = ''

                # profile_link 칼럼 추가
                influencer_counts['profile_link'] = influencer_counts['username'].map(profile_links)
                
                # Reels Views를 숫자로 변환하여 정렬에 사용
                def convert_reels_views(value):
                    try:
                        if isinstance(value, (int, float)):
                            return float(value)
                        if isinstance(value, str):
                            value = value.replace(',', '')
                            if 'K' in value:
                                return float(value.replace('K', '')) * 1000
                            elif 'M' in value:
                                return float(value.replace('M', '')) * 1000000
                            elif value.replace('.', '').isdigit():
                                return float(value)
                    except (ValueError, TypeError):
                        pass
                    return 0

                # Reels Views 숫자 변환
                influencer_counts['reels_views_numeric'] = influencer_counts['reels_views'].apply(convert_reels_views)
                
                # 브랜드 진행 횟수 내림차순, Reels Views 내림차순으로 정렬
                influencer_counts = influencer_counts.sort_values(
                    ['브랜드_진행_횟수', 'reels_views_numeric'], 
                    ascending=[False, False]
                )
                
                # 최종 컬럼 구성 및 이름 변경
                influencer_counts = influencer_counts[[
                    'username', 
                    'clean_name', 
                    '브랜드_진행_횟수', 
                    'brand_list',
                    'profile_link',
                    'reels_views'
                ]]
                influencer_counts.columns = [
                    'username',
                    'Clean Name',
                    '브랜드 진행 횟수',
                    '브랜드 목록',
                    '프로필 링크',
                    'Reels Views'
                ]
                
                # Sheet3에 저장
                influencer_counts.to_excel(writer, sheet_name='Sheet3', index=False)
                worksheet3 = writer.sheets['Sheet3']
                worksheet3.auto_filter.ref = worksheet3.dimensions
                
                # Sheet3의 프로필 링크 컬럼(5번째 열)을 하이퍼링크로 설정
                for row_idx, row in enumerate(worksheet3.iter_rows(min_row=2)):  # 헤더 제외
                    try:
                        link_cell = row[4]  # E열 (프로필 링크)
                        if link_cell.value and isinstance(link_cell.value, str) and (link_cell.value.startswith('http://') or link_cell.value.startswith('https://')):
                            link_cell.hyperlink = link_cell.value
                            link_cell.style = 'Hyperlink'
                    except Exception:
                        continue
                
                for column in worksheet3.columns:
                    worksheet3.column_dimensions[column[0].column_letter].width = 25

                # 타겟 유저의 Reels Views 값 가져오기 (숫자가 아닌 경우 0으로 처리)
                target_reels_views = 0
                try:
                    if isinstance(user_reels_views, (int, float)):
                        target_reels_views = float(user_reels_views)
                    elif isinstance(user_reels_views, str) and user_reels_views.replace('.', '').isdigit():
                        target_reels_views = float(user_reels_views)
                except (ValueError, TypeError):
                    pass

                # 회색 배경색 스타일 정의
                grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                # 각 행을 순회하면서 Reels Views 비교
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):  # 헤더 제외
                    try:
                        reels_views_cell = row[6]  # G열 (인덱스 6)
                        if reels_views_cell.value and str(reels_views_cell.value).replace('.', '').replace('K', '000').replace('M', '000000').isdigit():
                            influencer_views = float(str(reels_views_cell.value).replace('K', '000').replace('M', '000000'))
                            if influencer_views > target_reels_views:
                                for cell in row:
                                    cell.fill = grey_fill
                    except (ValueError, TypeError):
                        continue

                # Sheet2 칼럼 너비 설정
                for column in worksheet.columns:
                    worksheet.column_dimensions[column[0].column_letter].width = 25
                print(f"\nSheet2에 {len(df_overlapping)}개의 브랜드-인플루언서 매칭 정보가 저장되었습니다.")
        
        print(f"\nExcel 파일이 생성되었습니다: {excel_filename}")
        print(f"Sheet0: {target_username}의 브랜드 목록")
        print("Sheet1: 다른 인플루언서들의 브랜드 목록")
        if not df_overlapping.empty:
            print(f"Sheet2: {target_username} 브랜드를 진행한 인플루언서 목록")

except Exception as e:
    print(f"오류 발생: {e}")

finally:
    client.close() 